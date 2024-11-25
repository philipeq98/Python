import os
import openpyxl
from translate import Translator

def translate_excel_files(input_dir, output_dir):
    print("Starting translation process...")
    
    # Ensure the output directory exists
    os.makedirs(output_dir, exist_ok=True)
    print(f"Output directory '{output_dir}' is ready.")
    
    # Initialize the translator
    translator = Translator(to_lang="en", from_lang="ja")
    print("Translator initialized.")
    
    # Iterate over all Excel files in the input directory
    for filename in os.listdir(input_dir):
        if filename.endswith(".xlsx"):
            print(f"Processing file: {filename}")
            # Load the workbook and iterate through sheets
            input_file_path = os.path.join(input_dir, filename)
            workbook = openpyxl.load_workbook(input_file_path)
            print(f"Loaded workbook: {filename}")
            
            for sheet_name in workbook.sheetnames:
                print(f"Processing sheet: {sheet_name}")
                sheet = workbook[sheet_name]
                
                # Iterate through each cell in the sheet
                for row in sheet.iter_rows():
                    for cell in row:
                        if cell.value:
                            # Translate the cell value if it's a string
                            if isinstance(cell.value, str):
                                try:
                                    translated_text = translator.translate(cell.value)
                                    cell.value = translated_text
                                except Exception as e:
                                    print(f"Error translating cell {cell.coordinate} in file {filename}: {e}")
            
            # Save the translated workbook to the output directory
            output_file_path = os.path.join(output_dir, filename)
            workbook.save(output_file_path)
            print(f"Translated {filename} and saved to {output_file_path}")
    
    print("Translation process completed.")

# Example usage
input_directory = r'C:\Users\NowocinskiFilip\Desktop\DATA ENCODED'
output_directory = r'C:\Users\NowocinskiFilip\Desktop\DATA TRANSLATED'
translate_excel_files(input_directory, output_directory)
